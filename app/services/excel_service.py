"""
Excel Import/Export Service - Master data import and report export.

Features:
- Import customers, products, accounts
- Export reports to Excel
- Template generation
"""
from decimal import Decimal
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO
import json

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from sqlalchemy.orm import Session

from app.database.models import Customer, Product, Account, generate_uuid


class ExcelService:
    """Service for Excel import/export operations."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def _check_openpyxl(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel operations. Install with: pip install openpyxl")
    
    # ==================== TEMPLATE GENERATION ====================
    
    def generate_import_template(self, template_type: str) -> bytes:
        """Generate Excel template for data import."""
        self._check_openpyxl()
        
        wb = Workbook()
        ws = wb.active
        
        if template_type == 'customers':
            ws.title = "Customers"
            headers = ['Name*', 'Trade Name', 'GSTIN', 'PAN', 'Email', 'Phone', 
                      'Contact Person', 'Billing Address Line 1', 'Billing City', 
                      'Billing State', 'Billing Pincode', 'Customer Type']
            ws.append(headers)
            # Add sample row
            ws.append(['ABC Company', 'ABC Trading', '27AAAAA0000A1Z5', 'AAAAA0000A',
                      'abc@example.com', '9876543210', 'John Doe', '123 Main St',
                      'Mumbai', 'Maharashtra', '400001', 'b2b'])
        
        elif template_type == 'products':
            ws.title = "Products"
            headers = ['Name*', 'SKU', 'Description', 'HSN Code', 'Unit Price*',
                      'Unit', 'GST Rate', 'Opening Stock', 'Min Stock Level',
                      'Is Service']
            ws.append(headers)
            ws.append(['Widget A', 'WA001', 'Premium widget', '84716090', '1000',
                      'unit', '18', '100', '10', 'FALSE'])
        
        elif template_type == 'accounts':
            ws.title = "Accounts"
            headers = ['Code*', 'Name*', 'Account Type*', 'Parent Code', 
                      'Opening Balance', 'Description']
            ws.append(headers)
            ws.append(['1010', 'Cash in Hand', 'asset', '1000', '50000', 'Petty cash'])
        
        elif template_type == 'opening_balances':
            ws.title = "Opening Balances"
            headers = ['Account Code*', 'Opening Balance*', 'Notes']
            ws.append(headers)
            ws.append(['1010', '50000', 'Cash opening balance'])
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    # ==================== IMPORT OPERATIONS ====================
    
    def import_customers(
        self,
        company_id: str,
        file_data: bytes,
    ) -> Dict:
        """Import customers from Excel file."""
        self._check_openpyxl()
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                customer = Customer(
                    id=generate_uuid(),
                    company_id=company_id,
                    name=str(row[0]),
                    trade_name=str(row[1]) if row[1] else None,
                    gstin=str(row[2]) if row[2] else None,
                    pan=str(row[3]) if row[3] else None,
                    email=str(row[4]) if row[4] else None,
                    phone=str(row[5]) if row[5] else None,
                    contact_person=str(row[6]) if row[6] else None,
                    billing_address_line1=str(row[7]) if row[7] else None,
                    billing_city=str(row[8]) if row[8] else None,
                    billing_state=str(row[9]) if row[9] else None,
                    billing_pincode=str(row[10]) if row[10] else None,
                    customer_type=str(row[11]).lower() if row[11] else 'b2c',
                    is_active=True,
                )
                
                self.db.add(customer)
                imported += 1
                
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        self.db.commit()
        
        return {
            'imported': imported,
            'errors': errors,
            'total_rows': row_num - 1,
        }
    
    def import_products(
        self,
        company_id: str,
        file_data: bytes,
    ) -> Dict:
        """Import products from Excel file."""
        self._check_openpyxl()
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            
            try:
                product = Product(
                    id=generate_uuid(),
                    company_id=company_id,
                    name=str(row[0]),
                    sku=str(row[1]) if row[1] else None,
                    description=str(row[2]) if row[2] else None,
                    hsn_code=str(row[3]) if row[3] else None,
                    unit_price=Decimal(str(row[4])) if row[4] else Decimal('0'),
                    unit=str(row[5]) if row[5] else 'unit',
                    gst_rate=str(row[6]) if row[6] else '18',
                    opening_stock=Decimal(str(row[7])) if row[7] else Decimal('0'),
                    current_stock=Decimal(str(row[7])) if row[7] else Decimal('0'),
                    min_stock_level=Decimal(str(row[8])) if row[8] else Decimal('0'),
                    is_service=str(row[9]).upper() == 'TRUE' if row[9] else False,
                    is_active=True,
                )
                
                self.db.add(product)
                imported += 1
                
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        self.db.commit()
        
        return {
            'imported': imported,
            'errors': errors,
            'total_rows': row_num - 1,
        }
    
    def import_accounts(
        self,
        company_id: str,
        file_data: bytes,
    ) -> Dict:
        """Import chart of accounts from Excel file."""
        self._check_openpyxl()
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:  # Skip empty rows or missing required fields
                continue
            
            try:
                # Look up parent account if provided
                parent_id = None
                if row[3]:  # Parent Code
                    parent = self.db.query(Account).filter(
                        Account.company_id == company_id,
                        Account.code == str(row[3])
                    ).first()
                    if parent:
                        parent_id = parent.id
                
                # Determine account type
                account_type_str = str(row[2]).lower() if row[2] else 'asset'
                from app.database.models import AccountType
                account_type = AccountType(account_type_str)
                
                account = Account(
                    id=generate_uuid(),
                    company_id=company_id,
                    code=str(row[0]),
                    name=str(row[1]),
                    account_type=account_type,
                    parent_id=parent_id,
                    description=str(row[5]) if row[5] else None,
                    is_active=True,
                )
                
                self.db.add(account)
                self.db.flush()
                
                # Create opening balance transaction if provided
                opening_balance = Decimal(str(row[4])) if row[4] else Decimal('0')
                if opening_balance != Decimal('0'):
                    from app.services.accounting_service import AccountingService
                    from app.database.models import Company
                    company = self.db.query(Company).filter(Company.id == company_id).first()
                    if company:
                        accounting_service = AccountingService(self.db)
                        accounting_service.create_opening_balance_transaction(
                            company=company,
                            account_id=account.id,
                            amount=opening_balance,
                        )
                
                imported += 1
                
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        self.db.commit()
        
        return {
            'imported': imported,
            'errors': errors,
            'total_rows': row_num - 1,
        }
    
    def import_opening_balances(
        self,
        company_id: str,
        file_data: bytes,
    ) -> Dict:
        """Import opening balances for accounts from Excel file."""
        self._check_openpyxl()
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        updated = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                # Find account by code
                account = self.db.query(Account).filter(
                    Account.company_id == company_id,
                    Account.code == str(row[0])
                ).first()
                
                if not account:
                    errors.append({'row': row_num, 'error': f"Account code '{row[0]}' not found"})
                    continue
                
                # Create opening balance transaction
                opening_balance = Decimal(str(row[1])) if row[1] else Decimal('0')
                if opening_balance != Decimal('0'):
                    from app.services.accounting_service import AccountingService
                    from app.database.models import Company
                    company = self.db.query(Company).filter(Company.id == company_id).first()
                    if company:
                        accounting_service = AccountingService(self.db)
                        # Get current balance to calculate adjustment
                        current_balance = accounting_service.get_account_balance(account.id)
                        diff = opening_balance - current_balance
                        if diff != Decimal('0'):
                            accounting_service.create_opening_balance_transaction(
                                company=company,
                                account_id=account.id,
                                amount=diff,
                            )
                
                updated += 1
                
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        self.db.commit()
        
        return {
            'imported': updated,
            'errors': errors,
            'total_rows': row_num - 1,
        }
    
    # ==================== EXPORT OPERATIONS ====================
    
    def export_to_excel(
        self,
        data: List[Dict],
        columns: List[str],
        title: str = "Report",
        include_totals: bool = False,
    ) -> bytes:
        """Export data to Excel file."""
        self._check_openpyxl()
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Add title
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add date
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])  # Empty row
        
        # Add headers
        ws.append(columns)
        header_row = 4
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        for col_num, cell in enumerate(ws[header_row], 1):
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        totals = {col: Decimal('0') for col in columns}
        
        for row in data:
            row_data = []
            for col in columns:
                value = row.get(col, '')
                row_data.append(value)
                
                # Accumulate numeric totals
                if include_totals and isinstance(value, (int, float, Decimal)):
                    totals[col] += Decimal(str(value))
            
            ws.append(row_data)
        
        # Add totals row
        if include_totals:
            ws.append([])
            total_row = ['Total' if i == 0 else (float(totals[col]) if isinstance(totals[col], Decimal) and totals[col] != 0 else '') for i, col in enumerate(columns)]
            ws.append(total_row)
            
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_report(
        self,
        report_data: Dict,
        report_type: str,
    ) -> bytes:
        """Export a report to Excel with proper formatting."""
        self._check_openpyxl()
        
        if report_type == 'ledger':
            return self._export_ledger_report(report_data)
        elif report_type == 'aging':
            return self._export_aging_report(report_data)
        elif report_type == 'trial_balance':
            return self._export_trial_balance(report_data)
        else:
            # Generic export
            entries = report_data.get('entries', report_data.get('details', []))
            if entries:
                columns = list(entries[0].keys()) if entries else []
                return self.export_to_excel(entries, columns, report_type.replace('_', ' ').title())
            return b''
    
    def _export_ledger_report(self, data: Dict) -> bytes:
        """Export ledger report with running balance."""
        columns = ['Date', 'Voucher No', 'Type', 'Description', 'Debit', 'Credit', 'Balance']
        
        rows = []
        for entry in data.get('entries', []):
            rows.append({
                'Date': entry.get('date'),
                'Voucher No': entry.get('voucher_number'),
                'Type': entry.get('voucher_type'),
                'Description': entry.get('description'),
                'Debit': entry.get('debit'),
                'Credit': entry.get('credit'),
                'Balance': entry.get('balance'),
            })
        
        return self.export_to_excel(
            rows, columns, 
            f"Ledger - {data.get('account', {}).get('name', 'Account')}",
            include_totals=True
        )
    
    def _export_aging_report(self, data: Dict) -> bytes:
        """Export aging report."""
        columns = ['Invoice No', 'Date', 'Due Date', 'Customer/Vendor', 'Total', 'Outstanding', 'Days Overdue', 'Bucket']
        
        rows = []
        for entry in data.get('details', []):
            rows.append({
                'Invoice No': entry.get('invoice_number'),
                'Date': entry.get('invoice_date'),
                'Due Date': entry.get('due_date'),
                'Customer/Vendor': entry.get('customer_name') or entry.get('vendor_name'),
                'Total': entry.get('total_amount'),
                'Outstanding': entry.get('outstanding'),
                'Days Overdue': entry.get('days_overdue'),
                'Bucket': entry.get('bucket'),
            })
        
        return self.export_to_excel(
            rows, columns,
            f"{data.get('report_type', 'Aging')} Report",
            include_totals=True
        )
    
    def _export_trial_balance(self, data: Dict) -> bytes:
        """Export trial balance."""
        columns = ['Account Code', 'Account Name', 'Type', 'Debit', 'Credit']
        
        rows = []
        for entry in data.get('accounts', []):
            rows.append({
                'Account Code': entry.get('code'),
                'Account Name': entry.get('name'),
                'Type': entry.get('type'),
                'Debit': entry.get('debit'),
                'Credit': entry.get('credit'),
            })
        
        return self.export_to_excel(rows, columns, "Trial Balance", include_totals=True)
